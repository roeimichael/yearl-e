"""Extract Maddison GDP + Brecke conflicts for a single year.

Inputs: data/raw/maddison.xlsx + data/raw/brecke.xlsx (run fetch_sources.py first).
Output: data/raw/{year}_extract.json — region-agnostic dump:
  {
    "year": 1719,
    "gdppc": {ISO3: {country, region, gdppc, pop}},
    "conflicts": [{name, sy, ey, region_code, fatalities}]
  }
"""
from __future__ import annotations
import json
import sys
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", category=UserWarning)

ROOT = Path(__file__).parent.parent
RAW = ROOT / "data" / "raw"


def maddison_for_year(year: int) -> dict:
    wb = openpyxl.load_workbook(RAW / "maddison.xlsx", data_only=True, read_only=True)
    ws = wb["Full data"]
    out = {}
    for r in ws.iter_rows(values_only=True):
        if r[0] == "countrycode" or r[3] != year:
            continue
        if r[4] is None:
            continue
        out[r[0]] = {"country": r[1], "region": r[2], "gdppc": float(r[4]), "pop": float(r[5] or 0)}
    wb.close()
    return out


def brecke_active_in(year: int) -> list[dict]:
    wb = openpyxl.load_workbook(RAW / "brecke.xlsx", data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    out = []
    for r in ws.iter_rows(values_only=True):
        sy, ey = r[6], r[9]
        if not (isinstance(sy, (int, float)) and isinstance(ey, (int, float))):
            continue
        if sy <= year <= ey:
            out.append({
                "name": r[1],
                "sy": int(sy),
                "ey": int(ey),
                "region_code": r[12],
                "fatalities": r[5] if isinstance(r[5], (int, float)) else None,
            })
    wb.close()
    return out


def main() -> int:
    if len(sys.argv) != 2:
        print("usage: python scripts/fetch_year.py <year>", file=sys.stderr)
        return 2
    year = int(sys.argv[1])
    gdppc = maddison_for_year(year)
    conflicts = brecke_active_in(year)
    out = {"year": year, "gdppc": gdppc, "conflicts": conflicts}
    p = RAW / f"{year}_extract.json"
    p.write_text(json.dumps(out, indent=2), encoding="utf-8")
    print(f"wrote {p} | gdppc={len(gdppc)} countries, conflicts={len(conflicts)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
