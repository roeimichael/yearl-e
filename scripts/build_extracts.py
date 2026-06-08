"""Bulk-build per-year extracts (Maddison GDP + Brecke conflicts) for a range,
opening each workbook ONCE instead of per-year (fetch_year reopens them every
call — fine for one year, far too slow for 200).

Output matches fetch_year.py exactly: data/raw/{year}_extract.json with
  {year, gdppc: {ISO3: {country, region, gdppc, pop}}, conflicts: [...]}

Brecke only runs 1400-2000; years past its last record get no Brecke conflicts
(wire UCDP separately for 2001+). Skips Wikipedia (era_summary is optional).

Usage: python scripts/build_extracts.py 1816 2000
       python scripts/build_extracts.py 1816 2000 --force
"""
from __future__ import annotations
import argparse
import json
import sys
import warnings
from collections import defaultdict
from pathlib import Path

import openpyxl

import hced_lookup
import ucdp_lookup

warnings.filterwarnings("ignore", category=UserWarning)

ROOT = Path(__file__).parent.parent
RAW = ROOT / "data" / "raw"

# Brecke's catalog starts here; before it, HCED supplies conflicts (kept in sync
# with rank_year.BRECKE_FIRST_YEAR / war_regime).
BRECKE_FIRST_YEAR = 1400


def maddison_by_year() -> dict[int, dict]:
    """{year: {ISO3: {country, region, gdppc, pop}}} from Maddison 'Full data',
    one streaming pass."""
    wb = openpyxl.load_workbook(RAW / "maddison.xlsx", data_only=True, read_only=True)
    ws = wb["Full data"]
    out: dict[int, dict] = defaultdict(dict)
    for r in ws.iter_rows(values_only=True):
        if r[0] == "countrycode" or not isinstance(r[3], (int, float)):
            continue
        if r[4] is None:
            continue
        out[int(r[3])][r[0]] = {
            "country": r[1], "region": r[2],
            "gdppc": float(r[4]), "pop": float(r[5] or 0),
        }
    wb.close()
    return out


def brecke_all() -> list[dict]:
    """Every Brecke conflict as {name, sy, ey, region_code, fatalities}, once."""
    wb = openpyxl.load_workbook(RAW / "brecke.xlsx", data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    out = []
    for r in ws.iter_rows(values_only=True):
        sy, ey = r[6], r[9]
        if not (isinstance(sy, (int, float)) and isinstance(ey, (int, float))):
            continue
        out.append({
            "name": r[1], "sy": int(sy), "ey": int(ey),
            "region_code": r[12],
            "fatalities": r[5] if isinstance(r[5], (int, float)) else None,
        })
    wb.close()
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("start", type=int)
    ap.add_argument("end", type=int)
    ap.add_argument("--force", action="store_true",
                    help="Rewrite extracts that already exist.")
    args = ap.parse_args()

    print(f"loading Maddison + Brecke once for {args.start}..{args.end} ...")
    madd = maddison_by_year()
    conflicts = brecke_all()
    brecke_max = max(c["ey"] for c in conflicts)
    ucdp_lo, ucdp_hi = ucdp_lookup.coverage()
    hced_lo, hced_hi = hced_lookup.coverage()
    print(f"  Maddison years: {min(madd)}..{max(madd)} | HCED: {hced_lo}..{hced_hi} "
          f"(used <{BRECKE_FIRST_YEAR}) | Brecke: {len(conflicts)} (ends {brecke_max}) "
          f"| UCDP: {ucdp_lo}..{ucdp_hi}")

    wrote = skipped = 0
    for year in range(args.start, args.end + 1):
        p = RAW / f"{year}_extract.json"
        if p.exists() and not args.force:
            skipped += 1
            continue
        # Each source owns its slice of the timeline (no double-count): HCED
        # before 1400, Brecke 1400-1999, UCDP from 2000 on.
        if year < BRECKE_FIRST_YEAR:
            active = hced_lookup.active_in(year)
        elif year <= brecke_max:
            active = [c for c in conflicts if c["sy"] <= year <= c["ey"]]
        else:
            active = ucdp_lookup.active_in(year)
        out = {"year": year, "gdppc": madd.get(year, {}), "conflicts": active}
        p.write_text(json.dumps(out, indent=2), encoding="utf-8", newline="\n")
        wrote += 1
    if args.end > ucdp_hi:
        print(f"  NOTE: years {ucdp_hi + 1}..{args.end} have no conflict data "
              f"(beyond both Brecke and UCDP) — safety at baseline there.")
    print(f"DONE wrote {wrote}, skipped {skipped}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
